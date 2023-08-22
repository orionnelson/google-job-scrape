# Fuctions of combine results
"""
:return: 
- Read all the CSV files from the folder.
- Combine them into a single DataFrame.
- Identify keywords that meet your criteria.
- Group the data by the identified keywords and remove duplicates.
- Write the grouped data to separate sheets in an Excel fil
"""
import os
import colorsys
import pandas as pd
import nltk
from nltk.corpus import stopwords
from collections import Counter
from openpyxl.styles import PatternFill

# Download the stopwords
nltk.download('stopwords')

def get_file_keywords(filename):
    return filename.replace('.csv', '').replace('_', ' ').split()

# Generates the Position Of Authority Terms to avoid using and pagnation headers
def generate_position_auth_terms():
    # Step 1: Define the base terms
    full_terms = ['junior', 'senior', 'principal', 'chair', 'manager', 'director','officer','engineer','architect', 'associate', 'developer', 'assistant', 'analyst']
    abbreviations = ['jr', 'sr','eng' ]

    # Step 2: Generate variations
    full_terms_variations = {term for word in full_terms for term in [word, word.upper(), word.capitalize()]}
    abbreviations_variations = {abbrev for abbrev_word in abbreviations for abbrev in [abbrev_word, abbrev_word.upper(), abbrev_word.capitalize()]}

    # Step 3: Generate variations with and without periods (for abbreviations)
    abbreviations_with_periods = {abbrev + '.' for abbrev in abbreviations_variations}

    # Combine all variations
    position_auth_terms = full_terms_variations.union(abbreviations_variations).union(abbreviations_with_periods)
    return position_auth_terms

def identify_keywords(filenames):
    file_name_mod_all_words = [get_file_keywords(filename) for filename in filenames]
    all_words =  [item for sublist in file_name_mod_all_words for item in sublist]
    word_count = Counter(all_words)
    
    stop_words = set(stopwords.words('english'))

    # Define position authority abbreviations and their variants
    position_auth_terms = generate_position_auth_terms()

    # Filter out common words, position authority terms, and get words with count > 1
    filtered_word_counts = {word: count for word, count in word_count.items() if count > 1 and word.lower() not in stop_words and word.lower() not in position_auth_terms}
    
    # Sort words by count in descending order
    sorted_uncommon_words = sorted(filtered_word_counts, key=filtered_word_counts.get, reverse=True)
    
    return sorted_uncommon_words

def group_files_by_keyword(filenames, keywords):
    grouped_files = {}
    assigned_files = set()

    # Sort keywords by occurrence, starting with the least common
    sorted_keywords = keywords

    for keyword in sorted_keywords:
        grouped_files[keyword] = []
        for file in filenames:
            if keyword in get_file_keywords(file) and file not in assigned_files:
                grouped_files[keyword].append(file)
                assigned_files.add(file)

    # Remove empty groups
    grouped_files = {k: v for k, v in grouped_files.items() if v}


    # Group the leftover files under their own names
    for file in filenames:
        if file not in assigned_files:
            group_name = ' '.join(get_file_keywords(file))
            grouped_files[group_name] = [file]
            assigned_files.add(file)
    # Get the leftovers from what we are searching. not in assigned_files and give it its own group based on the file name
    # ' '.join(get_file_keywords(filename))


    return grouped_files

# Auto Generate Colors Based on the amount of diffrent entries
def generate_colors(n):
    hues = [i/n for i in range(n)]
    colors = [colorsys.hls_to_rgb(hue, 0.7, 0.6) for hue in hues]  # 0.7 lightness and 0.6 saturation
    hex_colors = ["".join(f"{int(255*x):02X}" for x in rgb) for rgb in colors]
    return hex_colors


def combine_csv_to_excel(folder_path):
    all_files = [f for f in os.listdir(folder_path) if f.endswith('.csv')]
    
    keywords = identify_keywords(all_files)
    file_groups = group_files_by_keyword(all_files, keywords)

    with pd.ExcelWriter('output.xlsx', engine='openpyxl') as writer:
        for keyword, files in file_groups.items():
            frames = []
            colors = generate_colors(len(files))
            colors_used = {}
            row_offsets = []

            for file in files:
                df = pd.read_csv(os.path.join(folder_path, file))
                df.drop_duplicates(inplace=True)
                frames.append(df)
                row_offsets.append(len(df))
            
            combined = pd.concat(frames, ignore_index=True)
            combined.drop_duplicates(inplace=True)
            combined.to_excel(writer, sheet_name=keyword, index=False, startrow=1)
            
            # Now, modify the cells for coloring
            current_offset = 3  # Considering a header row
            for idx, length in enumerate(row_offsets):
                for row in range(current_offset, current_offset + length):
                    for col in range(1, combined.shape[1] + 1):
                        cell = writer.sheets[keyword].cell(row=row, column=col)
                        cell.fill = PatternFill(start_color=colors[idx], end_color=colors[idx], fill_type="solid")
                current_offset += length
                colors_used[files[idx]] = colors[idx]

            # Adding legend at the end
            legend_start_row = current_offset + 1
            for idx, (file, color) in enumerate(colors_used.items()):
                legend_cell = writer.sheets[keyword].cell(row=legend_start_row + idx, column=1)
                legend_cell.value = file.replace('.csv', '')
                legend_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    return 'output.xlsx'



if __name__ == '__main__':
    combine_csv_to_excel(os.path.join(os.path.dirname('__file__'),'results'))



